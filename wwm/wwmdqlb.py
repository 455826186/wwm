from openpyxl import load_workbook
from openpyxl import Workbook


def login_1():
    data_2 = []  # 定义一个列表
    wb = load_workbook('qcd_cs.xlsx')
    sheet = wb['login']
    for i in range(2, sheet.max_row+1):
        case = []
        for j in range(1, sheet.max_column+1):  # 用for循环取出表格中的值
            print(sheet.cell(row=i, column=j).value)  # 按照顺序输出值
            case.append(sheet.cell(row=i, column=j).value)
        print(case)
        data_2.append(case)
    print(data_2)
    return data_2
    print('返回的所有数据为:',data_2)

login_1()
